import pandas as pd
import glob
import os
import re
from copy import copy
from datetime import datetime
import openpyxl

def adjust_formula(formula_str, prev_row, new_row):
    """
    Adjusts cell reference row numbers in Excel formulas.
    E.g., '=I804' becomes '=I805' and 'C804' becomes 'C805'.
    """
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return formula_str
    
    # Matches letters followed by the old row number
    pattern = rf'([A-Za-z]+){prev_row}\b'
    return re.sub(pattern, rf'\g<1>{new_row}', formula_str)


def add_row_to_excel_table(file_path, sheet_name, table_name, new_data, output_path=None):
    """
    Appends a new record to a formal Excel Table, preserving formatting,
    auto-updating formulas, and expanding the Table's range boundaries.
    """
    if output_path is None:
        output_path = file_path

    # Load workbook preserving macros/VBA
    wb = openpyxl.load_workbook(file_path, keep_vba=True)
    ws = wb[sheet_name]
    
    # Get the Excel Table
    table = ws.tables.get(table_name)
    if not table:
        raise ValueError(f"Table '{table_name}' not found in sheet '{sheet_name}'.")

    # Parse table range (e.g., 'B2:W804')
    match = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', table.ref)
    if not match:
        raise ValueError(f"Unsupported table reference format: {table.ref}")
        
    start_col_letter, start_row_str, end_col_letter, end_row_str = match.groups()
    start_row = int(start_row_str)
    last_row = int(end_row_str)
    new_row = last_row + 1

    # Map column letters and headers
    start_col_idx = openpyxl.utils.column_index_from_string(start_col_letter)
    end_col_idx = openpyxl.utils.column_index_from_string(end_col_letter)
    
    # Create header to column index map (e.g. {'DESDE': 2, 'VENCIMIENTO': 3, ...})
    header_mapping = {}
    for col in range(start_col_idx, end_col_idx + 1):
        header_val = ws.cell(row=start_row, column=col).value
        if header_val:
            header_mapping[header_val] = col

    print(f"Adding new row {new_row} to Table '{table_name}'...")

    # Write data and copy styles/formulas
    for col in range(start_col_idx, end_col_idx + 1):
        header_val = ws.cell(row=start_row, column=col).value
        prev_cell = ws.cell(row=last_row, column=col)
        new_cell = ws.cell(row=new_row, column=col)
        
        # 1. Copy Style from the row above
        if prev_cell.has_style:
            new_cell.font = copy(prev_cell.font)
            new_cell.border = copy(prev_cell.border)
            new_cell.fill = copy(prev_cell.fill)
            new_cell.number_format = copy(prev_cell.number_format)
            new_cell.protection = copy(prev_cell.protection)
            new_cell.alignment = copy(prev_cell.alignment)

        # 2. Write Value or Formula
        if header_val in new_data:
            # Write literal value from user dictionary
            new_cell.value = new_data[header_val]
        else:
            # Check if row above had a formula; if so, copy and adjust it
            prev_val = prev_cell.value
            if isinstance(prev_val, str) and prev_val.startswith('='):
                adjusted = adjust_formula(prev_val, last_row, new_row)
                new_cell.value = adjusted
            else:
                # No value provided, and no formula in row above
                new_cell.value = None

    # 3. Expand Table range reference to include the new row
    new_table_ref = f"{start_col_letter}{start_row}:{end_col_letter}{new_row}"
    table.ref = new_table_ref
    print(f"Table range updated from {start_col_letter}{start_row}:{end_col_letter}{last_row} to {new_table_ref}")

    # Save workbook
    wb.save(output_path)
    print(f"Workbook successfully saved to: {output_path}")



def read_csv_data(csv_file):
    """
    Reads the CSV file with semicolon delimiter
    """
    # Try different encodings
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            return pd.read_csv(csv_file, delimiter=';', encoding=encoding)
        except UnicodeDecodeError:
            continue
    
    # If none worked, try with error handling
    return pd.read_csv(csv_file, delimiter=';', encoding='utf-8', errors='ignore')

def filter_records(df):
    """
    Filters records based on the conditions:
    - tipo_operacion = COMPRA
    - tipo_documento = NOTA_CREDITO or similar
    """
    # Filter for COMPRA operations
    df_filtered = df[df['tipo_operacion'] == 'COMPRA'].copy()
    
    # Filter for NOTA_CREDITO or similar documents
    df_filtered = df_filtered[df_filtered['tipo_documento'].str.contains('NOTA_CREDITO', case=False, na=False)]
    
    return df_filtered

def prepare_titulos_activos_data(row, today_date):
    """
    Prepares data for TitulosActivos8 table
    """
    return {
        "Título": "Nota Crédito SRI",
        "Titular": row['propietario'],
        "Fecha de \nCompra": pd.to_datetime(today_date),
        "Valor \nNominal": float(row['valor_nominal']),
        "% Precio Comprado": float(row['precio']),
        "% Precio Neto Comprado": float(row['precio_neto']),
        "Comisión Santa Fé\nCompra": float(row['comision_operador']),
        "Comisión\nBolsa Compra": float(row['total_comisiones']) - float(row['comision_operador']),
        "Operación Compra No.": int(row['operacion_no']) if str(row['operacion_no']).replace('.', '').replace('-', '').isdigit() else row['operacion_no'],
        "Estado": "Activo"
    }

def prepare_operaciones_reales_data(row, today_date, valor_comprado_con_comision):
    """
    Prepares data for BaseOperacionesReales table
    """
    return {
        "Fecha": pd.to_datetime(today_date),
        "Titular": row['propietario'],
        "Operación": "Compra Nota Crédito",
        "Valor": -abs(valor_comprado_con_comision)  # Negative value
    }

def get_valor_comprado_con_comision(titulos_df, new_row_data):
    """
    Calculates the Valor Comprado Con Comisión for the new record
    This is typically: Valor Nominal * (% Precio Neto Comprado / 100) + Total Comisiones
    """
    valor_nominal = new_row_data["Valor \nNominal"]
    precio_neto = new_row_data["% Precio Neto Comprado"]
    total_comisiones = new_row_data["Comisión Santa Fé\nCompra"] + new_row_data["Comisión\nBolsa Compra"]
    
    valor_comprado_sin_comision = valor_nominal * (precio_neto / 100)
    valor_comprado_con_comision = valor_comprado_sin_comision + total_comisiones
    
    return valor_comprado_con_comision

def main():
    # Find the Excel file
    pattern = "INV_Inversiones*.xlsm"
    excel_files = glob.glob(pattern)
    if not excel_files:
        print(f"No Excel file matching '{pattern}' found.")
        return
    
    excel_file = excel_files[0]
    print(f"Using Excel file: {excel_file}")
    
    # Find the CSV file
    csv_pattern = "resultados_pdf_*.csv"
    csv_files = glob.glob(csv_pattern)
    if not csv_files:
        print(f"No CSV file matching '{csv_pattern}' found.")
        return
    
    csv_file = csv_files[0]
    print(f"Using CSV file: {csv_file}")
    
    # Read CSV data
    try:
        df = read_csv_data(csv_file)
        print(f"CSV loaded successfully. Shape: {df.shape}")
    except Exception as e:
        print(f"Error reading CSV: {e}")
        return
    
    # Filter records based on conditions
    filtered_df = filter_records(df)
    print(f"Found {len(filtered_df)} records matching the criteria")
    
    if len(filtered_df) == 0:
        print("No records to process. Exiting.")
        return
    
    # Get today's date
    today_date = datetime.now().strftime("%Y-%m-%d")
    print(f"Using today's date: {today_date}")
    
    # Create output file name
    dir_name = os.path.dirname(excel_file)
    base_name = os.path.basename(excel_file)
    name_part, ext_part = os.path.splitext(base_name)
    output_file = os.path.join(dir_name, f"{name_part}_updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext_part}")
    
    # Copy the original file to output location first
    import shutil
    shutil.copy2(excel_file, output_file)
    print(f"Created backup copy: {output_file}")
    
    try:
        # Process each filtered record
        for index, row in filtered_df.iterrows():
            print(f"\nProcessing record {index + 1}:")
            print(f"  Propietario: {row['propietario']}")
            print(f"  Operación No: {row['operacion_no']}")
            print(f"  Valor Nominal: {row['valor_nominal']}")
            
            # Prepare data for TitulosActivos8 table
            titulos_data = prepare_titulos_activos_data(row, today_date)
            
            # Don't set Valor Comprado Con Comisión - let Excel calculate it with its formula
            
            # Add row to TitulosActivos8 table
            add_row_to_excel_table(
                file_path=output_file,
                sheet_name="Base Títulos",
                table_name="TitulosActivos8",
                new_data=titulos_data,
                output_path=output_file
            )
            
            # We need to get the calculated value from Excel for BaseOperacionesReales
            # For now, let's calculate it manually for the operaciones table
            valor_nominal = float(row['valor_nominal'])
            precio_neto = float(row['precio_neto'])
            total_comisiones = float(row['total_comisiones'])
            valor_comprado_con_comision = float(row['total_comprador_neto'])
            
            # Prepare data for BaseOperacionesReales table
            operaciones_data = prepare_operaciones_reales_data(row, today_date, valor_comprado_con_comision)
            
            # Add row to BaseOperacionesReales table
            add_row_to_excel_table(
                file_path=output_file,
                sheet_name="Base OperacionesReales",
                table_name="BaseOperacionesReales",
                new_data=operaciones_data,
                output_path=output_file
            )
            
            print(f"  [OK] Successfully added records to both tables")
        
        print(f"\n[OK] All records processed successfully!")
        print(f"Updated file saved as: {output_file}")
        
    except Exception as e:
        print(f"Error processing records: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
