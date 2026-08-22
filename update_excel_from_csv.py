import pandas as pd
import glob
import os
import re
from copy import copy
from datetime import datetime
import openpyxl

def adjust_formula(formula_str, prev_row, new_row):
    """
    Adjusts cell reference row numbers in Excel formulas.
    E.g., '=I804' becomes '=I805' and 'C804' becomes 'C805'.
    """
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return formula_str
    
    # Matches letters followed by the old row number
    pattern = rf'([A-Za-z]+){prev_row}\b'
    return re.sub(pattern, rf'\g<1>{new_row}', formula_str)


def add_row_to_excel_table(file_path, sheet_name, table_name, new_data, output_path=None):
    """
    Appends a new record to a formal Excel Table, preserving formatting,
    auto-updating formulas, and expanding the Table's range boundaries.
    """
    if output_path is None:
        output_path = file_path

    # Load workbook preserving macros/VBA
    wb = openpyxl.load_workbook(file_path, keep_vba=True)
    ws = wb[sheet_name]
    
    # Get the Excel Table
    table = ws.tables.get(table_name)
    if not table:
        raise ValueError(f"Table '{table_name}' not found in sheet '{sheet_name}'.")

    # Parse table range (e.g., 'B2:W804')
    match = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', table.ref)
    if not match:
        raise ValueError(f"Unsupported table reference format: {table.ref}")
        
    start_col_letter, start_row_str, end_col_letter, end_row_str = match.groups()
    start_row = int(start_row_str)
    last_row = int(end_row_str)
    new_row = last_row + 1

    # Map column letters and headers
    start_col_idx = openpyxl.utils.column_index_from_string(start_col_letter)
    end_col_idx = openpyxl.utils.column_index_from_string(end_col_letter)
    
    # Create header to column index map (e.g. {'DESDE': 2, 'VENCIMIENTO': 3, ...})
    header_mapping = {}
    for col in range(start_col_idx, end_col_idx + 1):
        header_val = ws.cell(row=start_row, column=col).value
        if header_val:
            header_mapping[header_val] = col

    print(f"Adding new row {new_row} to Table '{table_name}'...")

    # Write data and copy styles/formulas
    for col in range(start_col_idx, end_col_idx + 1):
        header_val = ws.cell(row=start_row, column=col).value
        prev_cell = ws.cell(row=last_row, column=col)
        new_cell = ws.cell(row=new_row, column=col)
        
        # 1. Copy Style from the row above
        if prev_cell.has_style:
            new_cell.font = copy(prev_cell.font)
            new_cell.border = copy(prev_cell.border)
            new_cell.fill = copy(prev_cell.fill)
            new_cell.number_format = copy(prev_cell.number_format)
            new_cell.protection = copy(prev_cell.protection)
            new_cell.alignment = copy(prev_cell.alignment)

        # 2. Write Value or Formula
        if header_val in new_data:
            # Write literal value from user dictionary
            new_cell.value = new_data[header_val]
        else:
            # Check if row above had a formula; if so, copy and adjust it
            prev_val = prev_cell.value
            if isinstance(prev_val, str) and prev_val.startswith('='):
                adjusted = adjust_formula(prev_val, last_row, new_row)
                new_cell.value = adjusted
            else:
                # No value provided, and no formula in row above
                new_cell.value = None

    # 3. Expand Table range reference to include the new row
    new_table_ref = f"{start_col_letter}{start_row}:{end_col_letter}{new_row}"
    table.ref = new_table_ref
    print(f"Table range updated from {start_col_letter}{start_row}:{end_col_letter}{last_row} to {new_table_ref}")

    # Save workbook
    wb.save(output_path)
    print(f"Workbook successfully saved to: {output_path}")



def read_csv_data(csv_file):
    """
    Reads the CSV file with semicolon delimiter
    """
    # Try different encodings
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            return pd.read_csv(csv_file, delimiter=';', encoding=encoding)
        except UnicodeDecodeError:
            continue
    
    # If none worked, try with error handling
    return pd.read_csv(csv_file, delimiter=';', encoding='utf-8', errors='ignore')

def filter_records(df):
    """
    Filters records based on the conditions:
    - tipo_operacion = COMPRA
    - tipo_documento = NOTA_CREDITO or similar
    """
    # Filter for COMPRA operations
    df_filtered = df[df['tipo_operacion'] == 'COMPRA'].copy()
    
    # Filter for NOTA_CREDITO or similar documents
    df_filtered = df_filtered[df_filtered['tipo_documento'].str.contains('NOTA_CREDITO', case=False, na=False)]
    
    return df_filtered

def prepare_titulos_activos_data(row, today_date):
    """
    Prepares data for TitulosActivos8 table
    """
    return {
        "Título": "Nota Crédito SRI",
        "Titular": row['propietario'],
        "Fecha de \nCompra": pd.to_datetime(today_date),
        "Valor \nNominal": float(row['valor_nominal']),
        "% Precio Comprado": float(row['precio']),
        "% Precio Neto Comprado": float(row['precio_neto']),
        "Comisión Santa Fé\nCompra": float(row['comision_operador']),
        "Comisión\nBolsa Compra": float(row['total_comisiones']) - float(row['comision_operador']),
        "Operación Compra No.": int(row['operacion_no']) if str(row['operacion_no']).replace('.', '').replace('-', '').isdigit() else row['operacion_no'],
        "Estado": "Activo"
    }

def prepare_operaciones_reales_data(row, today_date, valor_comprado_con_comision):
    """
    Prepares data for BaseOperacionesReales table
    """
    return {
        "Fecha": pd.to_datetime(today_date),
        "Titular": row['propietario'],
        "Operación": "Compra Nota Crédito",
        "Valor": -abs(valor_comprado_con_comision)  # Negative value
    }

def get_valor_comprado_con_comision(titulos_df, new_row_data):
    """
    Calculates the Valor Comprado Con Comisión for the new record
    This is typically: Valor Nominal * (% Precio Neto Comprado / 100) + Total Comisiones
    """
    valor_nominal = new_row_data["Valor \nNominal"]
    precio_neto = new_row_data["% Precio Neto Comprado"]
    total_comisiones = new_row_data["Comisión Santa Fé\nCompra"] + new_row_data["Comisión\nBolsa Compra"]
    
    valor_comprado_sin_comision = valor_nominal * (precio_neto / 100)
    valor_comprado_con_comision = valor_comprado_sin_comision + total_comisiones
    
    return valor_comprado_con_comision

import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import shutil

CONFIG_FILE = "config.json"

DEFAULT_CONFIG_PATHS = {
    "carpetaOrigenExcel": "C:/Users/super/DATOS/100_COMPARTIDO/Inversiones/",
    "carpetaOrigenCsv": "C:/PROYECTOS/DescargaDiaria/Salida/",
    "carpetaSalida": "C:/Users/super/DATOS/100_COMPARTIDO/Inversiones/"
}

def load_config_paths():
    """Carga las rutas de configuración desde config.json."""
    if not os.path.exists(CONFIG_FILE):
        return DEFAULT_CONFIG_PATHS.copy()
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        paths = data.get("paths", {})
        return {
            "carpetaOrigenExcel": paths.get("carpetaOrigenExcel", DEFAULT_CONFIG_PATHS["carpetaOrigenExcel"]),
            "carpetaOrigenCsv": paths.get("carpetaOrigenCsv", DEFAULT_CONFIG_PATHS["carpetaOrigenCsv"]),
            "carpetaSalida": paths.get("carpetaSalida", DEFAULT_CONFIG_PATHS["carpetaSalida"]),
        }
    except Exception as e:
        print(f"Error al leer config.json: {e}")
        return DEFAULT_CONFIG_PATHS.copy()


def save_config_paths(new_paths):
    """Guarda o actualiza la sección paths en config.json respetando otras claves existentes."""
    data = {}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = {}
    
    if "paths" not in data:
        data["paths"] = {}
    
    data["paths"]["carpetaOrigenExcel"] = new_paths.get("carpetaOrigenExcel", "")
    data["paths"]["carpetaOrigenCsv"] = new_paths.get("carpetaOrigenCsv", "")
    data["paths"]["carpetaSalida"] = new_paths.get("carpetaSalida", "")
    
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error al guardar config.json: {e}")
        return False


def process_excel_update(excel_file, csv_file, output_dir=None, log_callback=None):
    """
    Procesa la actualización del archivo Excel a partir del CSV.
    Permite especificar una carpeta de salida y una función callback para registros de log.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"No se encontró el archivo Excel: {excel_file}")
    
    if not os.path.exists(csv_file):
        raise FileNotFoundError(f"No se encontró el archivo CSV: {csv_file}")

    log(f"Cargando archivo Excel: {excel_file}")
    log(f"Cargando archivo CSV: {csv_file}")

    # Cargar datos CSV
    df = read_csv_data(csv_file)
    log(f"CSV cargado correctamente. Registros totales: {len(df)}")

    # Filtrar registros
    filtered_df = filter_records(df)
    log(f"Registros coincidentes con el filtro (COMPRA / NOTA_CREDITO): {len(filtered_df)}")

    if len(filtered_df) == 0:
        log("No hay registros para procesar.")
        return None

    today_date = datetime.now().strftime("%Y-%m-%d")

    # Definir directorio de salida
    if not output_dir:
        output_dir = os.path.dirname(excel_file) or "."
    
    os.makedirs(output_dir, exist_ok=True)

    base_name = os.path.basename(excel_file)
    name_part, ext_part = os.path.splitext(base_name)
    output_file = os.path.join(output_dir, f"{name_part}_updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext_part}")

    # Crear copia del archivo original
    shutil.copy2(excel_file, output_file)
    log(f"Copia de respaldo creada en: {output_file}")

    # Procesar registros
    for index, row in filtered_df.iterrows():
        log(f"\nProcesando registro {index + 1}:")
        log(f"  Propietario: {row['propietario']}")
        log(f"  Operación No: {row['operacion_no']}")
        log(f"  Valor Nominal: {row['valor_nominal']}")

        # Datos para TitulosActivos8
        titulos_data = prepare_titulos_activos_data(row, today_date)
        add_row_to_excel_table(
            file_path=output_file,
            sheet_name="Base Títulos",
            table_name="TitulosActivos8",
            new_data=titulos_data,
            output_path=output_file
        )

        # Datos para BaseOperacionesReales
        valor_comprado_con_comision = float(row['total_comprador_neto'])
        operaciones_data = prepare_operaciones_reales_data(row, today_date, valor_comprado_con_comision)

        add_row_to_excel_table(
            file_path=output_file,
            sheet_name="Base OperacionesReales",
            table_name="BaseOperacionesReales",
            new_data=operaciones_data,
            output_path=output_file
        )
        log("  [OK] Registro añadido exitosamente a ambas tablas.")

    log(f"\n[OK] ¡Proceso finalizado con éxito!")
    log(f"Archivo actualizado guardado en: {output_file}")
    return output_file


def launch_gui():
    """Lanza la interfaz gráfica Tkinter para seleccionar archivos, carpetas y configurar rutas."""
    root = tk.Tk()
    root.title("Actualizador de Excel desde CSV")
    root.geometry("800x620")

    # Contenedor con Pestañas
    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True, padx=8, pady=5)

    tab_procesamiento = ttk.Frame(notebook, padding="10")
    tab_configuracion = ttk.Frame(notebook, padding="10")

    notebook.add(tab_procesamiento, text="Procesamiento")
    notebook.add(tab_configuracion, text="Configuración de Rutas")

    # Variables de Procesamiento
    excel_var = tk.StringVar()
    csv_var = tk.StringVar()
    output_var = tk.StringVar()

    # Variables de Configuración
    cfg_excel_dir_var = tk.StringVar()
    cfg_csv_dir_var = tk.StringVar()
    cfg_output_dir_var = tk.StringVar()

    def auto_detect_files():
        """Auto-detecta los archivos Excel y CSV basándose en las carpetas configuradas."""
        cfg_paths = load_config_paths()
        
        # Cargar valores en las variables de la pestaña de configuración
        cfg_excel_dir_var.set(cfg_paths["carpetaOrigenExcel"])
        cfg_csv_dir_var.set(cfg_paths["carpetaOrigenCsv"])
        cfg_output_dir_var.set(cfg_paths["carpetaSalida"])

        # Buscar Excel en carpetaOrigenExcel o directorio actual
        excel_dir = cfg_paths["carpetaOrigenExcel"]
        excel_found = []
        if os.path.exists(excel_dir):
            excel_found = glob.glob(os.path.join(excel_dir, "INV_Inversiones*.xlsm"))
            if not excel_found:
                excel_found = glob.glob(os.path.join(excel_dir, "*.xlsm"))
        if not excel_found:
            excel_found = glob.glob("INV_Inversiones*.xlsm")

        if excel_found:
            excel_var.set(os.path.abspath(excel_found[0]))

        # Buscar CSV en carpetaOrigenCsv o directorio actual
        csv_dir = cfg_paths["carpetaOrigenCsv"]
        csv_found = []
        if os.path.exists(csv_dir):
            csv_found = glob.glob(os.path.join(csv_dir, "resultados_pdf_*.csv"))
            if not csv_found:
                csv_found = glob.glob(os.path.join(csv_dir, "*.csv"))
        if not csv_found:
            csv_found = glob.glob("resultados_pdf_*.csv")

        if csv_found:
            csv_var.set(os.path.abspath(csv_found[0]))

        # Carpeta de salida
        out_dir = cfg_paths["carpetaSalida"]
        if out_dir and os.path.exists(out_dir):
            output_var.set(os.path.abspath(out_dir))
        elif excel_var.get():
            output_var.set(os.path.dirname(excel_var.get()))
        else:
            output_var.set(os.getcwd())

    # ==================== PESTAÑA 1: PROCESAMIENTO ====================
    file_frame = ttk.LabelFrame(tab_procesamiento, text=" Selección de Archivos de Origen ", padding="10")
    file_frame.pack(fill="x", pady=5)
    file_frame.columnconfigure(1, weight=1)

    # 1. Archivo Excel
    ttk.Label(file_frame, text="Archivo Excel (.xlsm):").grid(row=0, column=0, sticky="w", pady=4)
    ttk.Entry(file_frame, textvariable=excel_var).grid(row=0, column=1, sticky="ew", padx=5, pady=4)
    def browse_excel():
        init_dir = cfg_excel_dir_var.get() if os.path.exists(cfg_excel_dir_var.get()) else "."
        path = filedialog.askopenfilename(initialdir=init_dir, filetypes=[("Excel Habilitado para Macros", "*.xlsm"), ("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")])
        if path:
            excel_var.set(path)
    ttk.Button(file_frame, text="Examinar...", command=browse_excel).grid(row=0, column=2, pady=4)

    # 2. Archivo CSV
    ttk.Label(file_frame, text="Archivo CSV (Resultados):").grid(row=1, column=0, sticky="w", pady=4)
    ttk.Entry(file_frame, textvariable=csv_var).grid(row=1, column=1, sticky="ew", padx=5, pady=4)
    def browse_csv():
        init_dir = cfg_csv_dir_var.get() if os.path.exists(cfg_csv_dir_var.get()) else "."
        path = filedialog.askopenfilename(initialdir=init_dir, filetypes=[("Archivos CSV", "*.csv"), ("Todos los archivos", "*.*")])
        if path:
            csv_var.set(path)
    ttk.Button(file_frame, text="Examinar...", command=browse_csv).grid(row=1, column=2, pady=4)

    # Consola de Registro (Log)
    log_frame = ttk.LabelFrame(tab_procesamiento, text=" Progreso de Ejecución ", padding="5")
    log_frame.pack(fill="both", expand=True, pady=5)

    log_text = tk.Text(log_frame, wrap="word", height=10, font=("Consolas", 9))
    log_text.pack(side="left", fill="both", expand=True)
    scrollbar = ttk.Scrollbar(log_frame, command=log_text.yview)
    scrollbar.pack(side="right", fill="y")
    log_text.config(yscrollcommand=scrollbar.set)

    def append_log(msg):
        log_text.insert("end", str(msg) + "\n")
        log_text.see("end")

    # Botones de Acción
    action_frame = ttk.Frame(tab_procesamiento)
    action_frame.pack(fill="x", pady=6)
    action_frame.columnconfigure(0, weight=1)
    action_frame.columnconfigure(1, weight=1)

    btn_run = ttk.Button(action_frame, text="🚀 Iniciar Procesamiento")
    btn_run.grid(row=0, column=0, sticky="ew", padx=(0, 4))

    last_generated_file = [None]

    def open_generated_excel():
        path = last_generated_file[0]
        if path and os.path.exists(path):
            try:
                os.startfile(os.path.normpath(path))
            except Exception as ex:
                messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{ex}")
        else:
            messagebox.showerror("Error", "No hay ningún archivo generado disponible para abrir.")

    btn_open_excel = ttk.Button(action_frame, text="📊 Abrir Excel Resultante", state="disabled", command=open_generated_excel)
    btn_open_excel.grid(row=0, column=1, sticky="ew", padx=(4, 0))

    def worker():
        btn_run.config(state="disabled")
        btn_open_excel.config(state="disabled")
        log_text.delete("1.0", "end")
        
        excel_path = excel_var.get().strip()
        csv_path = csv_var.get().strip()
        out_dir = cfg_output_dir_var.get().strip() or load_config_paths()["carpetaSalida"]

        try:
            res_file = process_excel_update(excel_path, csv_path, out_dir, log_callback=append_log)
            if res_file:
                last_generated_file[0] = res_file
                btn_open_excel.config(state="normal")
                messagebox.showinfo("Éxito", f"Procesamiento completado correctamente.\n\nArchivo generado:\n{res_file}")
            else:
                messagebox.showwarning("Sin Registros", "No se encontraron registros que cumplan las condiciones para procesar.")
        except Exception as ex:
            append_log(f"\n[ERROR] Ocurrió un fallo: {ex}")
            messagebox.showerror("Error", f"Ocurrió un error durante el procesamiento:\n{ex}")
        finally:
            btn_run.config(state="normal")

    def start_processing():
        if not excel_var.get().strip():
            messagebox.showerror("Campo Requerido", "Por favor seleccione el archivo Excel de origen.")
            return
        if not csv_var.get().strip():
            messagebox.showerror("Campo Requerido", "Por favor seleccione el archivo CSV de origen.")
            return
        
        threading.Thread(target=worker, daemon=True).start()

    btn_run.config(command=start_processing)

    # ==================== PESTAÑA 2: CONFIGURACIÓN DE RUTAS ====================
    cfg_frame = ttk.LabelFrame(tab_configuracion, text=" Configuración de Carpetas por Defecto ", padding="15")
    cfg_frame.pack(fill="x", pady=10)
    cfg_frame.columnconfigure(1, weight=1)

    ttk.Label(cfg_frame, text="Carpeta Origen Excel:").grid(row=0, column=0, sticky="w", pady=6)
    ttk.Entry(cfg_frame, textvariable=cfg_excel_dir_var).grid(row=0, column=1, sticky="ew", padx=5, pady=6)
    def browse_cfg_excel():
        path = filedialog.askdirectory(initialdir=cfg_excel_dir_var.get() or ".")
        if path:
            cfg_excel_dir_var.set(path)
    ttk.Button(cfg_frame, text="Examinar...", command=browse_cfg_excel).grid(row=0, column=2, pady=6)

    ttk.Label(cfg_frame, text="Carpeta Origen CSV:").grid(row=1, column=0, sticky="w", pady=6)
    ttk.Entry(cfg_frame, textvariable=cfg_csv_dir_var).grid(row=1, column=1, sticky="ew", padx=5, pady=6)
    def browse_cfg_csv():
        path = filedialog.askdirectory(initialdir=cfg_csv_dir_var.get() or ".")
        if path:
            cfg_csv_dir_var.set(path)
    ttk.Button(cfg_frame, text="Examinar...", command=browse_cfg_csv).grid(row=1, column=2, pady=6)

    ttk.Label(cfg_frame, text="Carpeta de Salida:").grid(row=2, column=0, sticky="w", pady=6)
    ttk.Entry(cfg_frame, textvariable=cfg_output_dir_var).grid(row=2, column=1, sticky="ew", padx=5, pady=6)
    def browse_cfg_output():
        path = filedialog.askdirectory(initialdir=cfg_output_dir_var.get() or ".")
        if path:
            cfg_output_dir_var.set(path)
    ttk.Button(cfg_frame, text="Examinar...", command=browse_cfg_output).grid(row=2, column=2, pady=6)

    def save_configuration_action():
        new_paths = {
            "carpetaOrigenExcel": cfg_excel_dir_var.get().strip(),
            "carpetaOrigenCsv": cfg_csv_dir_var.get().strip(),
            "carpetaSalida": cfg_output_dir_var.get().strip()
        }
        if save_config_paths(new_paths):
            auto_detect_files()
            messagebox.showinfo("Éxito", "Configuración guardada correctamente en config.json.")
        else:
            messagebox.showerror("Error", "No se pudo guardar la configuración en config.json.")

    btn_save_cfg = ttk.Button(tab_configuracion, text="💾 Guardar Configuración", command=save_configuration_action)
    btn_save_cfg.pack(fill="x", pady=15)

    # Cargar archivos iniciales según la configuración
    auto_detect_files()

    root.mainloop()


if __name__ == "__main__":
    launch_gui()


